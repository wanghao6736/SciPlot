"""
数据处理器具体实现模块，提供了表格文件处理器的实现。
支持CSV和Excel格式的文件读取、处理和验证。
"""
from typing import Any, Dict, Optional

import pandas as pd

from data_processing.base import FileProcessor
from data_processing.config import FileFormat, TableConfig
from data_processing.exceptions import ProcessorError, ValidationError
from data_processing.utils.formatters import DataFormatter


class TableFileProcessor(FileProcessor):
    """表格文件处理器
    
    实现了对CSV和Excel格式表格文件的处理功能，支持数据的读取、
    预处理、验证和后处理等操作。可以通过配置控制数据处理的各个环节。
    
    Attributes:
        config (TableConfig): 表格处理配置对象
        _formatter (Optional[DataFormatter]): 数据格式化器，可选
        
    Example:
        ```python
        # 创建配置
        config = TableConfig(
            input_path="data.csv",
            header_row=0,
            parse_dates=["date_column"]
        )
        
        # 创建处理器
        processor = TableFileProcessor(config)
        
        # 处理数据
        processor.process()
        
        # 检查数据是否有效
        if processor.validate():
            # 输出标准格式
            result = processor.to_standard_format()
            print(result)
        ```
    """
    
    def __init__(self, config: TableConfig, formatter: Optional[DataFormatter] = None):
        """初始化表格文件处理器
        
        Args:
            config: 表格处理配置
            formatter: 数据格式化器，可选
        """
        super().__init__(config)
        self._formatter = formatter
        self._data = None  # 处理后的数据
        self._raw_data = None  # 原始数据
        self._validated = False  # 是否已验证
        
        # 确保配置类型正确
        if not isinstance(config, TableConfig):
            raise TypeError(f"配置必须是TableConfig类型，而不是{type(config)}")
        
        # 验证配置
        self._validate_config()
    
    def _validate_config(self) -> None:
        """验证配置
        
        检查配置的有效性，包括文件格式、工作表名称等。
        
        Raises:
            ValidationError: 配置无效时抛出
        """
        # 基本配置验证在父类中已完成
        super()._validate_config()
        
        # 验证表格特有的配置
        if isinstance(self.config, TableConfig):
            # 工作表验证
            if self.config.file_format == FileFormat.EXCEL:
                if not self.config.sheet_name and not self.config.sheet_names:
                    # 使用默认工作表（第一个）
                    pass
                elif self.config.sheet_name and self.config.sheet_names:
                    # 不能同时指定sheet_name和sheet_names
                    raise ValidationError("不能同时指定sheet_name和sheet_names")
    
    def process(self) -> 'TableFileProcessor':
        """处理数据
        
        执行整个数据处理流程：读取文件、预处理、后处理。
        使用方法链式调用方式，便于链式处理。
        
        Returns:
            TableFileProcessor: self，便于链式调用
            
        Raises:
            ProcessorError: 处理错误时抛出
        """
        try:
            # 重置验证状态
            self._validated = False
            
            # 1. 读取文件
            if self.config.sheet_names:
                # 如果指定了多个工作表，处理所有工作表
                sheets_data = self.process_sheets()
                # 合并所有工作表数据
                self._raw_data = pd.concat(sheets_data.values())
            else:
                # 处理单个文件/工作表
                self._raw_data = self._read_file()
            
            # 检查数据是否为空
            if self._raw_data.empty:
                raise ProcessorError("读取的数据为空")
            
            # 2. 预处理
            preprocessed_data = self._preprocess(self._raw_data)
            
            # 3. 后处理
            self._data = self._postprocess(preprocessed_data)
            
            return self
        except Exception as e:
            raise ProcessorError(f"数据处理失败: {str(e)}") from e
    
    def process_sheets(self) -> Dict[str, pd.DataFrame]:
        """处理多个工作表
        
        读取并处理Excel文件中的多个工作表。
        
        Returns:
            Dict[str, pd.DataFrame]: 工作表名到数据框的映射
            
        Raises:
            ProcessorError: 处理错误时抛出
        """
        if self.config.file_format != FileFormat.EXCEL:
            raise ProcessorError("只有Excel文件支持多工作表处理")
        
        if not self.config.sheet_names:
            raise ProcessorError("未指定要处理的工作表名称")
        
        result = {}
        for sheet_name in self.config.sheet_names:
            # 暂时更改配置中的工作表名
            original_sheet_name = self.config.sheet_name
            self.config.sheet_name = sheet_name
            
            try:
                # 读取并处理单个工作表
                df = self._read_file()
                result[sheet_name] = df
            except Exception as e:
                raise ProcessorError(f"处理工作表 {sheet_name} 失败: {str(e)}") from e
            finally:
                # 恢复原始配置
                self.config.sheet_name = original_sheet_name
        
        return result
    
    def _build_read_params(self) -> Dict[str, Any]:
        """构建读取参数
        
        根据配置和文件格式构建相应的文件读取参数。
        为不同的读取方法(CSV、Excel)生成正确的参数集。
        
        Returns:
            Dict[str, Any]: 读取参数字典
        """
        # 通用参数(适用于所有格式)
        common_params = {}
        
        # 添加CSV/Excel共有的参数
        if self.config.header_row is not None:
            common_params["header"] = self.config.header_row
        if self.config.index_col is not None:
            common_params["index_col"] = self.config.index_col
        
        # 添加数据类型参数(对所有格式有效)
        if self.config.dtype is not None:
            common_params["dtype"] = self.config.dtype
        if self.config.parse_dates is not None:
            common_params["parse_dates"] = self.config.parse_dates
        
        # 根据文件格式构建特定参数
        if self.config.file_format == FileFormat.CSV:
            # CSV特有参数
            csv_params = {
                **common_params,
                "encoding": self.config.file_encoding,  # 只有CSV支持编码参数
            }
            
            # 添加数据范围参数
            if self.config.start_row is not None:
                csv_params["skiprows"] = self.config.start_row
                
            return csv_params
            
        elif self.config.file_format == FileFormat.EXCEL:
            # Excel特有参数(pandas.read_excel)
            excel_params = {
                **common_params,
                "engine": "openpyxl",  # 使用openpyxl引擎
            }
            
            # 添加工作表名称
            if self.config.sheet_name is not None:
                excel_params["sheet_name"] = self.config.sheet_name
                
            # Excel支持的行范围参数(不同于CSV)
            if self.config.start_row is not None:
                # Excel用skiprows接受整数或整数列表
                excel_params["skiprows"] = self.config.start_row
            
            return excel_params
            
        else:
            raise ProcessorError(f"不支持的文件格式: {self.config.file_format}")
    
    def _get_openpyxl_params(self) -> Dict[str, Any]:
        """获取openpyxl特定的参数
        
        构建直接使用openpyxl读取Excel文件的参数。
        
        Returns:
            Dict[str, Any]: openpyxl参数字典
        """
        # openpyxl的load_workbook参数
        return {
            "read_only": self.config.read_only,  # 只读模式提高大文件性能
            "data_only": True,  # 只读取值而非公式
        }
    
    def _read_file(self) -> pd.DataFrame:
        """读取文件
        
        根据配置读取文件内容。支持一次性读取或分块读取。
        
        Returns:
            pd.DataFrame: 读取的数据框
            
        Raises:
            ProcessorError: 读取错误时抛出
        """
        try:
            if self.config.chunk_size is not None and self.config.chunk_size > 0:
                # 分块读取大文件
                return self._read_file_in_chunks()
            else:
                # 一次性读取
                return self._read_file_at_once()
        except Exception as e:
            raise ProcessorError(f"读取文件失败: {str(e)}") from e
    
    def _read_file_at_once(self) -> pd.DataFrame:
        """一次性读取文件
        
        将整个文件内容一次性读入内存。
        
        Returns:
            pd.DataFrame: 读取的数据框
        """
        if self.config.file_format == FileFormat.CSV:
            # 读取CSV文件
            params = self._build_read_params()
            df = pd.read_csv(self.config.input_path, **params)
        elif self.config.file_format == FileFormat.EXCEL:
            # 读取Excel文件
            params = self._build_read_params()
            df = pd.read_excel(self.config.input_path, **params)
        else:
            raise ProcessorError(f"不支持的文件格式: {self.config.file_format}")
        
        # 应用行列范围限制
        return self._slice_data(df)
    
    def _read_file_in_chunks(self) -> pd.DataFrame:
        """
        分块读取文件，能够处理大文件
        
        Returns:
            pd.DataFrame: 处理后的数据框
        """
        chunks = []
        chunksize = self.config.chunk_size if self.config.chunk_size else 10000
        
        # 基于文件格式选择读取方法
        if self.config.file_format == FileFormat.CSV:
            # 构建读取参数
            pd_params = self._build_read_params()
            # 确保chunksize参数存在
            pd_params['chunksize'] = chunksize
            
            # 分块读取CSV
            for chunk in pd.read_csv(self.config.input_path, **pd_params):
                chunk = self._slice_data(chunk)
                chunks.append(chunk)
        
        elif self.config.file_format == FileFormat.EXCEL:
            try:
                # 获取配置的列范围
                start_col = self.config.start_col if self.config.start_col is not None else 1
                end_col = self.config.end_col if self.config.end_col is not None else None
                min_col = start_col
                max_col = end_col if end_col is not None else None
                
                # 获取openpyxl参数
                xl_params = self._get_openpyxl_params()
                import openpyxl
                
                # 打开工作簿
                wb = openpyxl.load_workbook(
                    self.config.input_path, 
                    **xl_params
                )
                
                # 选择工作表
                sheet_name = self.config.sheet_name
                if sheet_name:
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                    else:
                        self._log_warning(f"工作表 '{sheet_name}' 不存在，使用第一个工作表")
                        ws = wb.active
                else:
                    ws = wb.active
                
                # 获取配置的行范围
                start_row = self.config.start_row if self.config.start_row is not None else 1
                end_row = self.config.end_row if self.config.end_row is not None else ws.max_row
                
                # 获取表头
                header_row = start_row
                headers = []
                for row in ws.iter_rows(
                    min_row=header_row, 
                    max_row=header_row,
                    min_col=min_col,
                    max_col=max_col,
                    values_only=True
                ):
                    headers = list(row)
                    break
                
                # 如果没有获取到有效表头，使用默认列名
                if not headers:
                    self._log_warning("未能获取有效表头，使用默认列名")
                    # 获取最大列数
                    max_cols = 0
                    for row in ws.iter_rows(
                        min_row=start_row+1, 
                        max_row=min(start_row+10, end_row),
                        min_col=min_col,
                        max_col=max_col,
                        values_only=True
                    ):
                        max_cols = max(max_cols, len([cell for cell in row if cell is not None]))
                    
                    headers = [f"Column{i+1}" for i in range(max_cols)]
                
                # 分块读取数据行
                current_row = start_row + 1  # 跳过表头行
                while current_row <= end_row:
                    chunk_end = min(current_row + chunksize - 1, end_row)
                    chunk_data = []
                    print(f"读取第{current_row}行到第{chunk_end}行")
                    
                    for row in ws.iter_rows(
                        min_row=current_row, 
                        max_row=chunk_end,
                        min_col=min_col,
                        max_col=max_col,
                        values_only=True
                    ):
                        # 只添加非空行
                        if any(cell is not None for cell in row):
                            # 确保行长度与表头一致
                            row_data = list(row)
                            if len(row_data) < len(headers):
                                # 填充短行
                                row_data.extend([None] * (len(headers) - len(row_data)))
                            elif len(row_data) > len(headers):
                                # 截断长行
                                row_data = row_data[:len(headers)]
                            chunk_data.append(row_data)
                    
                    # 创建当前块的DataFrame
                    if chunk_data:
                        df_chunk = pd.DataFrame(chunk_data, columns=headers)
                        chunks.append(df_chunk)
                    
                    # 移动到下一个块
                    current_row = chunk_end + 1
                
                # 关闭工作簿
                wb.close()
            
            except (ImportError, Exception) as e:
                # 导入错误或其他异常，回退到pandas的默认读取方式
                error_type = "openpyxl不可用" if isinstance(e, ImportError) else f"使用openpyxl读取Excel失败: {str(e)}"
                self._log_warning(f"{error_type}，使用pandas读取Excel（效率较低）")
                
                pd_params = self._build_read_params()
                df = pd.read_excel(self.config.input_path, **pd_params)
                df = self._slice_data(df)
                
                # 分块处理
                total_rows = len(df)
                for i in range(0, total_rows, chunksize):
                    end = min(i + chunksize, total_rows)
                    chunks.append(df.iloc[i:end])
        
        else:
            raise ValueError(f"不支持的文件格式: {self.config.file_format}")
        
        # 将所有块合并为一个DataFrame
        if not chunks:
            return pd.DataFrame()
        
        return pd.concat(chunks, ignore_index=True)
    
    def _slice_data(self, df: pd.DataFrame) -> pd.DataFrame:
        """切片数据
        
        根据配置的行列范围限制切片数据框。
        
        Args:
            df: 要切片的数据框
            
        Returns:
            pd.DataFrame: 切片后的数据框
        """
        # 限制行范围
        if self.config.end_row is not None:
            # 需要调整以考虑header_row
            adjusted_end = self.config.end_row
            if self.config.header_row is not None:
                adjusted_end -= self.config.header_row
            df = df.iloc[:adjusted_end]
        
        # 限制列范围
        return self._slice_columns(df)
    
    def _slice_columns(self, df: pd.DataFrame) -> pd.DataFrame:
        """切片列
        
        根据配置的列范围限制切片数据框的列。
        
        Args:
            df: 要切片的数据框
            
        Returns:
            pd.DataFrame: 切片后的数据框
        """
        if self.config.start_col is not None and self.config.end_col is not None:
            # 使用列索引切片
            return df.iloc[:, self.config.start_col:self.config.end_col + 1]
        elif self.config.start_col is not None:
            # 只有起始列
            return df.iloc[:, self.config.start_col:]
        elif self.config.end_col is not None:
            # 只有结束列
            return df.iloc[:, :self.config.end_col + 1]
        else:
            # 没有列范围限制
            return df
    
    def _preprocess(self, df: pd.DataFrame) -> pd.DataFrame:
        """预处理
        
        执行数据预处理操作，如清洗、转换等。
        
        Args:
            df: 要预处理的数据框
            
        Returns:
            pd.DataFrame: 预处理后的数据框
        """
        # 去除空行
        df = df.dropna(how='all')
        
        # 去除全为空格的行
        if not df.empty:
            mask = df.apply(lambda x: x.astype(str).str.isspace().all(), axis=1)
            df = df[~mask]
        
        # 列名处理：去除空格
        df.columns = df.columns.str.strip()
        
        # 数据值处理：去除前后空格
        # 只处理字符串列
        for col in df.select_dtypes(include=['object']).columns:
            df[col] = df[col].astype(str).str.strip()
        
        return df
    
    def _postprocess(self, df: pd.DataFrame) -> pd.DataFrame:
        """后处理
        
        执行数据后处理操作，如格式化、计算派生指标等。
        
        Args:
            df: 要后处理的数据框
            
        Returns:
            pd.DataFrame: 后处理后的数据框
        """
        # 确保所有数值列使用数值类型
        for col in df.columns:
            # 尝试转换为数值类型
            try:
                # 如果列包含百分比字符串，先替换百分号
                if df[col].dtype == 'object':
                    df[col] = df[col].astype(str).replace('%', '', regex=True).replace(',', '', regex=True)
                # 转换为数值，错误标记为NaN
                df[col] = pd.to_numeric(df[col], errors='coerce')
            except Exception:
                pass  # 无法转换则保持原样
        
        return df
    
    def validate(self) -> bool:
        """验证数据
        
        检查处理后的数据是否符合验证规则。
        
        Returns:
            bool: 数据是否有效
        """
        if self._data is None:
            return False
        
        # 检查是否为空
        if self._data.empty:
            return False
        
        # 应用自定义验证规则
        if 'required_columns' in self.config.validation_rules:
            required_cols = self.config.validation_rules['required_columns']
            missing_cols = [col for col in required_cols if col not in self._data.columns]
            if missing_cols:
                return False
        
        # 标记为已验证
        self._validated = True
        return True
    
    def to_standard_format(self) -> Dict[str, Any]:
        """转换为标准格式
        
        将处理后的数据转换为标准字典格式，便于后续处理。
        
        标准格式为：
        {
            "data": {
                "column_name": [...],  # 列数据，一维数组
                ...
            },
            "metadata": {
                "rows": int,
                "columns": [...],
                "dtypes": {...}
            }
        }
        
        Returns:
            Dict[str, Any]: 标准格式的数据字典
            
        Raises:
            ProcessorError: 数据未处理时抛出
        """
        if self._data is None:
            raise ProcessorError("数据尚未处理")
        
        # 基础数据格式
        base_format = {
            "data": {
                col: self._data[col].values.tolist()
                for col in self._data.columns
            },
            "metadata": {
                "rows": len(self._data),
                "columns": list(self._data.columns),
                "dtypes": {
                    col: str(dtype)
                    for col, dtype in self._data.dtypes.items()
                }
            }
        }
        
        # 如果有格式化器，进行特定格式转换
        if self._formatter:
            try:
                return self._formatter.format(base_format)
            except Exception as e:
                raise ProcessorError(f"格式转换错误: {str(e)}")
        
        return base_format
    
    def _log_warning(self, message: str) -> None:
        """记录警告信息
        
        用于记录非致命性错误和警告信息。
        
        Args:
            message: 警告消息
        """
        # 这里可以实现更复杂的日志记录，目前只打印到控制台
        print(f"警告: {message}")
